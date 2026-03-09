#!/usr/bin/env python3
"""
Script to import employee data from Excel file to database
"""

import openpyxl
import sys
import os
from app import app, Employee, db

def import_employees_from_excel(excel_file_path):
    """Import employees from Excel file using openpyxl only"""
    
    try:
        # Read Excel file with openpyxl
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        
        # Get header row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value.lower() if cell.value else '')
        
        # Check if required columns exist
        required_columns = ['name', 'employee_id', 'department']
        for col in required_columns:
            if col not in headers:
                print(f"❌ Error: Column '{col}' not found in Excel file")
                print(f"Required columns: {required_columns}")
                print(f"Found columns: {headers}")
                return False
        
        # Get column indices
        name_col = headers.index('name')
        emp_id_col = headers.index('employee_id')
        dept_col = headers.index('department')
        
        with app.app_context():
            imported_count = 0
            skipped_count = 0
            
            # Skip header row, start from row 2
            for row_num in range(2, sheet.max_row + 1):
                try:
                    # Get cell values
                    name = str(sheet.cell(row=row_num, column=name_col + 1).value or '').strip()
                    emp_id = str(sheet.cell(row=row_num, column=emp_id_col + 1).value or '').strip()
                    department = str(sheet.cell(row=row_num, column=dept_col + 1).value or '').strip()
                    
                    # Skip empty rows
                    if not name or not emp_id:
                        print(f"⚠️  Skipping empty row {row_num}")
                        continue
                    
                    # Check if employee already exists
                    existing_employee = Employee.query.filter_by(employee_id=emp_id).first()
                    
                    if existing_employee:
                        print(f"⚠️  Skipping existing employee: {name} ({emp_id})")
                        skipped_count += 1
                        continue
                    
                    # Create new employee
                    new_employee = Employee(
                        name=name,
                        employee_id=emp_id,
                        department=department
                    )
                    
                    db.session.add(new_employee)
                    imported_count += 1
                    print(f"✅ Imported: {name} ({emp_id}) - {department}")
                    
                except Exception as e:
                    print(f"❌ Error importing row {row_num}: {e}")
                    continue
            
            # Commit all changes
            db.session.commit()
            
            print(f"\n" + "="*50)
            print(f"📊 Import Summary:")
            print(f"✅ Successfully imported: {imported_count} employees")
            print(f"⚠️  Skipped (already exists): {skipped_count} employees")
            print(f"📋 Total processed: {imported_count + skipped_count}")
            
            return True
            
    except FileNotFoundError:
        print(f"❌ Error: File '{excel_file_path}' not found")
        return False
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return False

def show_current_employees():
    """Show current employees in database"""
    with app.app_context():
        employees = Employee.query.all()
        print(f"\n📋 Current employees in database ({len(employees)}):")
        for emp in employees:
            print(f"  - {emp.name} ({emp.employee_id}) - {emp.department}")

def create_sample_excel():
    """Create a sample Excel file template using openpyxl"""
    try:
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"
        
        # Add headers
        headers = ['name', 'employee_id', 'department']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Add sample data
        sample_data = [
            ['สมชาย ใจดี', 'EMP001', 'IT'],
            ['สมศรี รักดี', 'EMP002', 'HR'],
            ['วิชัย มั่งคั่ง', 'EMP003', 'Sales'],
            ['มานี แสนสุข', 'EMP004', 'Marketing'],
            ['ประสิทธิ์ โชคดี', 'EMP005', 'IT']
        ]
        
        for row_num, data in enumerate(sample_data, 2):
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        
        # Save file
        sample_file = 'Employee_List_Template.xlsx'
        wb.save(sample_file)
        print(f"✅ Sample template created: {sample_file}")
        print("You can use this template to create your employee list")
        
    except Exception as e:
        print(f"❌ Error creating sample file: {e}")

if __name__ == "__main__":
    print("🚀 Employee Import Tool (openpyxl version)")
    print("="*50)
    
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python import_employees.py <excel_file_path>    # Import from Excel")
        print("  python import_employees.py --show             # Show current employees")
        print("  python import_employees.py --template         # Create sample template")
        sys.exit(1)
    
    arg = sys.argv[1]
    
    if arg == "--show":
        show_current_employees()
    elif arg == "--template":
        create_sample_excel()
    elif arg.endswith('.xlsx') or arg.endswith('.xls'):
        if os.path.exists(arg):
            import_employees_from_excel(arg)
        else:
            print(f"❌ File '{arg}' not found")
    else:
        print("❌ Invalid argument. Use --show, --template, or provide Excel file path")
